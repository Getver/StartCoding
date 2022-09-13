# openpyxl : 파이썬에서 엑셀을 쉽게 다룰 수 있도록 도와주는 라이브러리
# pip install openpyxl

import openpyxl

# 엑셀 만들기
wb = openpyxl.Workbook()

# 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 엑셀 저장하기
wb.save(r'C:\startcode_python\05_주식\참가자_data.xlsx')

# 폴더 경로 복사, r은 안에 있는 모든 \가 문자열로써 취급되게 도와줌