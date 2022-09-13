# 과제

# 주식 현재가 크롤링했던 데이터를 아래 양식의 엑셀을 불러와서 B2, B3, B4에 저장

# 네이버 증권 사이트 현재가 데이터를 파이썬으로 가져온다
# 엑셀을 불러와서 데이터를 저장한다

import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r'C:\startcode_python\05_주식\참가자_data.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['주식']

codes = [
    '005930',
    '000660',
    '035720'
]

names = [
    '삼성전자',
    'sk하이닉스',
    '카카오'
]

for code in range(0, 3):
    a = 'B' + str(code + 1)
    url = "https://finance.naver.com/item/sise.naver?code=" + codes[code]
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one("#_nowVal").text
    price = price.replace(',', '')
    ws[a] = price

for name in range(0, 3):
    b = 'A' + str(name + 1)
    ws[b] = names[name]

wb.save(fpath)